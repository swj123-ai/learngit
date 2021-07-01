from openpyxl import load_workbook
import xlrd
import pprint
import json
import openpyxl

PATH = "/home/swj/auto-master/Data/data.xlsx"
SHEET_NAME = "Sheet1"
book = xlrd.open_workbook(PATH)
table = book.sheet_by_name(SHEET_NAME)


def operactionexcle(): #将excel中的数据存入字典
    book = xlrd.open_workbook(PATH)
    table = book.sheet_by_name(SHEET_NAME)
    row_num = table.nrows
    col_num = table.ncols
    #print(row_num,col_num)
    s = []
    key = table.row_values(0)
    #print("key=",key)
    if row_num <= 1:
        print("没数据")
    else:
        j =1
        for i in range(row_num-1):
            d = {}
            values = table.row_values(j)
            #print("values=",values)
            for x in range(col_num):
                d[key[x]]=values[x]
            j += 1
            s.append(d)
            #print(type(s))
            #print(s)
        for item in s:
            res_data = item['data']
            res_url = item['url']
            #print(res_data,res_url)
    return s,res_data,res_url

def write_result(rows,columns,value): #将测试结果保存进excel文档
    table.put_cell(rows=rows,columns=columns,value=value)








if __name__ == "__main__":
    operactionexcle()









# class HandleExcel:
#     def __init__(self,file_path,sheet_name):
#         self.file_name = file_path
#         self.work_book = load_workbook(filename=self.file_name)
#         self.sheet = self.work_book[sheet_name]
#
#     def get_all_data(self):
#         all_data = self.sheet.iter_rows(values_only=True)
#         title = all_data[0]
#         case_data = all_data[1:]
#         return title,all_data
#
#     def get_case_data_dict(self):
#         case_list = []
#         title,case_data = self.get_all_data()
#         for value in case_data:
#             result = dict(zip(title,value))
#             case_list.append(result)
#         return case_list
#
#     def write_result(self, rows, columns, result=None):
#         self.sheet.cell(row=rows, column=columns).value = result
#         # 保存
#         self.work_book.save()
#
# if __name__ == "__main__":
#     handle_excel = HandleExcel(file_path="/home/swj/auto-master/Data/data.xlsx",sheet_name="Sheet1")
#     case_datas = handle_excel.get_case_data_dict()
#     print(case_datas)
#     for case in case_datas:
#         res = login_test.Test_testlogin().user_login(url=case['url'],data=ast.literal_eval(case['expected_data']))
#         print(res)
